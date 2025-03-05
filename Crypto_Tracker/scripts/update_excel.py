import requests
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from fpdf import FPDF
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def authenticate_google_sheets(credentials_file, sheet_name):
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(creds)
    return client.open(sheet_name).sheet1

def fetch_crypto_data():
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {"vs_currency": "usd", "order": "market_cap_desc", "per_page": 50, "page": 1, "sparkline": False}
    response = requests.get(url, params=params)
    data = response.json()
    return pd.DataFrame([{ "Name": coin["name"], "Symbol": coin["symbol"].upper(), "Price": coin["current_price"],
                            "Market Cap": coin["market_cap"], "24h Change %": coin["price_change_percentage_24h"]} for coin in data])

def update_excel(df, filename="crypto_data.xlsx"):
    df.to_excel(filename, index=False, sheet_name="Crypto Market Data")
    
    wb = load_workbook(filename)
    ws = wb.active
    
    # Apply formatting
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column_letter  # Get column name
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col].width = max_length + 2
    
    wb.save(filename)
    print(f"✅ Excel file '{filename}' updated and structured successfully!")

def update_google_sheet(sheet, df):
    sheet.clear()
    sheet.append_row(df.columns.tolist())
    for row in df.values.tolist():
        sheet.append_row(row)
    print("✅ Google Sheet updated successfully!")

def generate_pdf(df, filename="crypto_report.pdf"):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, "Cryptocurrency Market Report", ln=True, align='C')
    pdf.ln(10)
    
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Top 5 Cryptocurrencies by Market Cap:", ln=True)
    pdf.set_font("Arial", "", 10)
    for i, row in df.nlargest(5, "Market Cap").iterrows():
        pdf.cell(0, 10, f"{row['Name']} ({row['Symbol']}): ${row['Market Cap']:,.0f}", ln=True)
    
    pdf.ln(5)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Highest & Lowest 24h Price Changes:", ln=True)
    pdf.set_font("Arial", "", 10)
    highest = df.loc[df["24h Change %"].idxmax()]
    lowest = df.loc[df["24h Change %"].idxmin()]
    pdf.cell(0, 10, f"Highest: {highest['Name']} ({highest['Symbol']}) {highest['24h Change %']:.2f}%", ln=True)
    pdf.cell(0, 10, f"Lowest: {lowest['Name']} ({lowest['Symbol']}) {lowest['24h Change %']:.2f}%", ln=True)
    
    pdf.output(filename)
    print(f"✅ PDF Report '{filename}' generated successfully!")

def main():
    SHEET_NAME = "Crypto_Tracker"
    CREDENTIALS_FILE = "credentials.json"
    
    sheet = authenticate_google_sheets(CREDENTIALS_FILE, SHEET_NAME)
    
    while True:
        df = fetch_crypto_data()
        update_excel(df)
        update_google_sheet(sheet, df)
        generate_pdf(df)
        
        print("🔄 Next update in 5 minutes...")
        time.sleep(300)

if __name__ == "__main__":
    main()
